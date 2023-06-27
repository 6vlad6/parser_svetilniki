# -*- coding: cp1251 -*-

import openpyxl
import json


workbook = openpyxl.load_workbook("result/eraworld/products_eraworld.xlsx")
sheet = workbook.active

with open("result/eraworld/data.json", 'r', encoding="utf-8") as file:
    products_str = file.read()
    products = json.loads(products_str)["products"]

    keys_places = {
        'Материал': 21,
        'Материал корпуса': 23,
        'Категория по ограничению яркости': 25,
        'Срок службы или срок годности': 27,
        'Датчик движения': 29,
        'Питающее напряжение, В': 31,
        'Срок службы, час': 33,
        'количество светодиодов': 35,
        'Класс защиты от поражения электрическим током': 37,
        'Бренд': 39,
        'Габаритная яркость': 41,
        'Страна производства': 43,
        'Коэффициент мощности, PF': 45,
        'Гарантия': 47,
        'Условия хранения': 49,
        'Преимущества': 51,
        'Диапазон рабочего напряжения в сети, В': 53,
        'Степень защиты от воздействия окружающей среды, IP': 55,
        'Тип светильника': 57,
        'Класс энергоэффективности': 59,
        'Диапазон регулировки высоты, см': 61,
        'Индекс цветопередачи': 63,
        'Модель': 65,
        'Тип рассеивателя': 67,
        'Коэффициент пульсации светового потока, %': 69,
        'Область применения': 71,
        'Частота сети, Гц': 73,
        'Температурные ограничения хранения и перевозки': 75,
        'Срок службы изделия': 77,
        'Цветовая температура, К': 79,
        'Количество светодиодов, шт': 81,
        'Наличие аллергенов и резких запахов': 83,
        'Класс светораспределения': 85,
        'Диапазон рабочих температур': 87,
        'Тип товара': 89,
        'Тип кривой силы света': 91,
        'Материал рассеивателя': 93,
        'Цвет': 95,
        'Наличие категории ЛВЖ и ГЖ': 97,
        'Источник света': 99,
        'Напряжение, В': 101,
        'Климатическое исполнение': 103,
        'Температура эксплуатации': 105,
        'Светоодача, Лм/Вт': 107,
        'Цвет корпуса': 109
    }

    i = 2
    keys = []
    for product in products:
        first_separate_key_index = 21
        first_separate_value_index = 22

        separate_keys = []
        for key in product.keys():

            if key == "Код товара":
                sheet[f'A{i}'] = str(product[key])

            elif key == "Название":
                sheet[f'B{i}'] = str(product[key])

            elif key == "Описание":
                sheet[f'D{i}'] = str(product[key]).strip()

            elif key == "Категория":
                sheet[f'P{i}'] = str(product[key])

            elif key == "Потребляемая мощность, Вт":
                sheet[f'F{i}'] = str(product[key])

            elif key == "Световой поток, Лм":
                sheet[f'H{i}'] = str(product[key])

            elif "Габариты" in str(key):
                try:
                    try:
                        gabs = str(product[key]).lower().split("x")

                        sheet[f'K{i}'] = str(gabs[0])
                        sheet[f'I{i}'] = str(gabs[1])
                        sheet[f'J{i}'] = str(gabs[2])
                    except:
                        gabs = str(product[key]).lower().split("х")

                        sheet[f'K{i}'] = str(gabs[0])
                        sheet[f'I{i}'] = str(gabs[1])
                        sheet[f'J{i}'] = str(gabs[2])
                except:
                    pass

            elif "Масса" in str(key):
                sheet[f'L{i}'] = float(str(product[key]).replace(",", "."))


            elif "Цена".lower() in key.lower() or key == "Артикул":
                pass

            else:
                separate_keys.append(key)

        for key in separate_keys:
            title = key
            if key == "Гарантийный срок, мес":
                title = 'Гарантия'

            cell_name = sheet.cell(row=i, column=keys_places[title], value=key)
            cell_value = sheet.cell(row=i, column=keys_places[title]+1, value=product[key])

        i += 1

    workbook.save("result/eraworld/products_eraworld.xlsx")
