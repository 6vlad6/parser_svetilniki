# -*- coding: cp1251 -*-

import openpyxl
import json


workbook = openpyxl.load_workbook("result/gcsvt/products_gcsvt.xlsx")
sheet = workbook.active

with open("result/gcsvt/data.json", 'r', encoding="utf-8") as file:
    products_str = file.read()
    products = json.loads(products_str)["products"]
    keys_places = {
        'КСС': 21,
        'Уровень шума, дБ': 23,
        'Индекс цветопередачи': 25,
        'Пульсации светового потока': 27,
        'Корпус': 29,
        'Гарантия': 31,
        'Ресурс работы светодиодов': 33,
        'Цветовая температура светодиода': 35,
        'Серия': 37,
        'Длина волны УФ лампы': 39,
        'Гарантия и блок питания': 41,
        'Пульсации светового потока, %': 43,
        'Димминг': 45,
        'IP': 47,
        'Срок службы лампы': 49,
        'Объем упаковки': 51,
        'Тип лампы': 53,
        'Для категорий помещений': 55,
        'Коэффициент мощности': 57,
        'Источник питания': 59,
        'БАП': 61,
        'Класс защиты от поражения эл. током': 63,
        'Эффективность, лм/Вт': 65,
        'Цвет корпуса': 67,
        'Срок службы светодиодов': 69,
        'Индекс цветопередачи и цветовая температура': 71,
        'Светодиоды': 73,
        'Дополнительные опции': 75,
        'Вторичная оптика': 77,
        'Рабочая температура окр. среды ': 79,
        'Мощность бактерицидного излучения ламп, Вт': 81,
        'Вторичная оптика/рассеиватель': 83,
        'Рабочий ток светодиода': 85,
        'Климатическое исполнение': 87,
        'Напряжение питания': 89,
        'Производительность, м3/час': 91,
        'Подключение питания': 93,
        'Счетчик наработки времени': 95,
        'Крепление': 97,
        'Рабочая температура окр. среды': 99,
        'Конфигуратор': 101,
    }
    i = 2
    codes = []
    for product in products:
        try:
            product_code = str(product["Код товара"])
            if product_code in codes:
                continue
            else:
                codes.append(product_code)
                sheet[f'A{i}'] = product_code
        except:
            pass


        separate_keys = []
        for key in product.keys():

            if key == "Название":
                sheet[f'B{i}'] = str(product[key])

            elif key == "Описание":
                sheet[f'D{i}'] = str(product[key])

            elif key == "Категория":
                sheet[f'P{i}'] = str(product[key])

            elif key == "Мощность, Вт":
                sheet[f'F{i}'] = float(product[key])

            elif key == "Световой поток, лм" or key == "Поток, лм":
                sheet[f'H{i}'] = float(product[key])

            elif "Габариты Д/Ш/В" in str(key):
                try:
                    try:
                        gabs = str(product[key]).split("x")

                        sheet[f'I{i}'] = float(gabs[1])
                        sheet[f'J{i}'] = float(gabs[2])
                        sheet[f'K{i}'] = float(gabs[0])
                    except:
                        gabs = str(product[key]).split("х")

                        sheet[f'I{i}'] = float(gabs[1])
                        sheet[f'J{i}'] = float(gabs[2])
                        sheet[f'K{i}'] = float(gabs[0])
                except:
                    pass

            elif "Масса" in str(key):
                sheet[f'L{i}'] = str(product[key])

            elif "Цена".lower() in key.lower() or key == "Артикул":
                pass

            else:
                separate_keys.append(key)
        try:
            separate_keys["Конфигуратор"] = separate_keys.pop("Конфигуратор")
        except:
            pass

        if "Конфигуратор" in separate_keys:
            options = product["Конфигуратор"]['Конфигурации']
            for option in options:
                if option['Название опции'] in separate_keys:
                    separate_keys.remove(option['Название опции'])

        for key in separate_keys:

            if key == "Конфигуратор":
                options = product[key]['Конфигурации']

                for option in options:
                    option_name = option["Название опции"]

                    if str(option["Название опции"]).strip() == 'Цветовая температура и индекс цветопередачи':
                        option_name = 'Индекс цветопередачи и цветовая температура'

                    elif str(option["Название опции"]).strip() == 'Рассеиватель':
                        option_name = 'Вторичная оптика/рассеиватель'

                    if option_name == "Удлинение кабеля":
                        cell_name = sheet.cell(row=i, column=103, value=option["Название опции"])
                        cell_value = sheet.cell(row=i, column=104,
                                                value="; ".join(option["Опции"]))

                    elif option_name == "Увеличение мощности":
                        cell_name = sheet.cell(row=i, column=105, value=option["Название опции"])
                        cell_value = sheet.cell(row=i, column=106,
                                                value="; ".join(option["Опции"]))

                    elif option_name == "Опции":
                        cell_name = sheet.cell(row=i, column=107, value=option["Название опции"])
                        cell_value = sheet.cell(row=i, column=108,
                                                value="; ".join(option["Опции"]))

                    elif option_name == "Степень защиты":
                        cell_name = sheet.cell(row=i, column=109, value=option["Название опции"])
                        cell_value = sheet.cell(row=i, column=110,
                                                value="; ".join(option["Опции"]))

                    else:
                        cell_name = sheet.cell(row=i, column=keys_places[option_name], value=option["Название опции"])
                        cell_value = sheet.cell(row=i, column=keys_places[option_name]+1,
                                                value="; ".join(option["Опции"]))

            else:
                if key == "Код товара":
                    continue
                else:
                    cell_name = sheet.cell(row=i, column=keys_places[key], value=key)
                    cell_value = sheet.cell(row=i, column=keys_places[key]+1, value=product[key])

        i += 1

    workbook.save("result/gcsvt/products_gcsvt.xlsx")

print(len(codes))
print(len(set(codes)))